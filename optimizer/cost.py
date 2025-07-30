from openpyxl import load_workbook # type: ignore
import xlwings as xw # type: ignore
import general_data as gd

def quit_excel():
    for app in xw.apps: app.quit()

def update_and_read_excel(filepath, part_entries, joint_entries=None, submodule_type=gd.WATER_COLLECTION_WELDED, part_start_row=4, joint_start_row=4, summary_row=2):
    """
    Update specific cells in an existing Excel file and read calculated values.
    
    Args:
        filepath: Path to Excel file
        part_entries: List of lists containing values to write to the Excel file
        joint_entries: List of lists containing values to write to the joint entries
        submodule_type: 'Water Collection Welded', 'Water Collection TriArmor', 'Water Collection Unwelded', or 'Water Distribution'
        part_start_row: Starting row for writing entries
        summary_row: Row for writing summary information

    Returns:
        List of calculated values from the Excel file
    """
    app = xw.App(visible=False)
    app.quit()
    
    part_sheet_name = 'BAC Part List'
    joint_sheet_name = 'Joints List'
    summary_sheet_name = 'Summary'

    def _clear_excel_contents():
        """
        Clear the contents of the specified sheets in the Excel file.
        """
        try:
            part_sheet = workbook[part_sheet_name]
            joint_sheet = workbook[joint_sheet_name]
            summary_sheet = workbook[summary_sheet_name]

            for row in part_sheet.iter_rows(min_row=part_start_row, max_col=14, max_row=part_sheet.max_row):
                for cell in row:
                    cell.value = None

            for row in joint_sheet.iter_rows(min_row=joint_start_row, max_col=3, max_row=joint_sheet.max_row):
                for cell in row:
                    cell.value = None

            for row in summary_sheet.iter_rows(min_row=summary_row, max_col=2, max_row=summary_sheet.max_row):
                for cell in row:
                    cell.value = None

            workbook.save(filepath)
        except Exception as e:
            print(f"Error clearing Excel contents: {e}")

    try:
        # Load the workbook and sheet
        workbook = load_workbook(filepath)
        _clear_excel_contents()

        part_sheet = workbook[part_sheet_name]

        # Write part entries to the Excel file
        for i, entry in enumerate(part_entries):
            for j, value in enumerate(entry):
                part_sheet.cell(row=part_start_row + i, column=j + 1, value=value)

        # Write joint entries to the Excel file
        if joint_entries is not None:
            joint_sheet = workbook[joint_sheet_name]
            for i, joint in enumerate(joint_entries):
                for j, value in enumerate(joint):
                    joint_sheet.cell(row=joint_start_row + i, column=j + 1, value=value)

        # Write submodule type to the summary sheet
        summary_sheet = workbook[summary_sheet_name]

        distinct_sets = sorted(list({entry[0] for entry in part_entries}))
        set_num = 0
        for i in range(summary_row, len(distinct_sets) + summary_row):
            summary_sheet.cell(row=i, column=1, value=distinct_sets[set_num])
            summary_sheet.cell(row=i, column=2, value=submodule_type)
            set_num += 1

        # Force recalculation of formulas
        workbook.save(filepath)
        _force_recalculation(filepath)

        # Reload the workbook to read calculated values
        workbook = load_workbook(filepath, data_only=True)
        summary_sheet = workbook[summary_sheet_name]
        
        # Read calculated values ()
        calculated_values = []
        for i in range(summary_row, len(distinct_sets) + summary_row):
            row_values = [summary_sheet.cell(row=i, column=col).value for col in range(1, 13)]  # Adjust columns as needed
            calculated_values.append(row_values)
        
        return calculated_values
    
    except Exception as e:
        print(f"Error updating or reading Excel: {e}")
        return None

def _force_recalculation(filepath):
    """
    Open the Excel file with xlwings to force recalculation of formulas.
    """
    try:
        app = xw.App(visible=False)
        wb = app.books.open(filepath)
        wb.app.calculate()  # Force recalculation
        wb.save()
        wb.close()
        app.quit()
    except Exception as e:
        print(f"Error forcing recalculation: {e}")

# Example usage:
# part_entries = [
#     ['Set1', 'ID1', 10, gd.CUT_TL, gd.FORM_RF, gd.GLV, 10, 5, 120.5, 3, 0, 50.0, 20.0, 'Class 1'],
#     ['Set1', 'ID2', 20, gd.CUT_MSP, gd.FORM_APB, gd.GLV, 14, 8, 200.0, 5, 4, 80.0, 30.0, 'Class 1']
# ]

# joint_entries = [['ID1:1', 'ID2:1', 15]]

# calculated_values = update_and_read_excel(
#     filepath='cost_calculator.xlsx',
#     part_entries=part_entries,
#     joint_entries=joint_entries
# )

# print("Calculated Values:", calculated_values)